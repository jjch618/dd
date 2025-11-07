import sys
import openpyxl
import matplotlib.pyplot as pyplot

def stoi(column):
	index = 0
	for char in column:
		index *= 26
		index += ord(char.upper()) - ord('A') + 1
	return index

def process_arguments():
	if len(sys.argv) < 3:
		print("Usage: python post-process.py <xlsx-file> <column-to-process>")
		sys.exit(1)
	
	if not sys.argv[1].endswith(".xlsx"):
		print("Error: File must be in .xlsx format")
		sys.exit(1)
	
	if not sys.argv[2].isalpha():
		print("Error: Column must be in alphabetic format")
		sys.exit(1)

	file = sys.argv[1]
	row = 4
	column = stoi(sys.argv[2])
	if (len(sys.argv) == 4):
		if not sys.argv[3].isdigit():
			print("Error: Row must be in numeric format")
			sys.exit(1)
		row = int(sys.argv[3])

	return file, row, column

def read_xlsx(file, row, column):
	data = []
	workbook = openpyxl.load_workbook(file)
	sheet = workbook.active
	for row in sheet.iter_rows(min_row=row, max_row=sheet.max_row, min_col=column, max_col=column):
		if row[0].value is None or row[0].value == "미제출":
			continue
		data.append(round(eval(row[0].value[1:])))
	workbook.close()
	return data

def plot_grade_distribution(data):
	pyplot.hist(data, bins=range(0, 101, 10), edgecolor='black')
	pyplot.xlabel("Score")
	pyplot.ylabel("Frequency")
	pyplot.title("Grade Distribution")
	pyplot.savefig("grade-distribution.png")
	pyplot.close()

def draw_grade_distribution(file, row, column):
	data = read_xlsx(file, row, column)
	plot_grade_distribution(data)

if __name__ == "__main__":
	draw_grade_distribution(*process_arguments())