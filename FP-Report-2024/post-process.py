import sys
import openpyxl

def stoi(column):
	index = 0
	for char in column:
		index *= 26
		index += ord(char.upper()) - ord('A') + 1
	return index

def process_arguments():
	if len(sys.argv) < 3:
		print("Usage: python post-process.py <xlsx-file> <column-to-process>")
		sys.exit(1)
	
	if not sys.argv[1].endswith(".xlsx"):
		print("Error: File must be in .xlsx format")
		sys.exit(1)
	
	if not sys.argv[2].isalpha():
		print("Error: Column must be in alphabetic format")
		sys.exit(1)

	file = sys.argv[1]
	row = 4
	column = stoi(sys.argv[2])
	if (len(sys.argv) == 4):
		if not sys.argv[3].isdigit():
			print("Error: Row must be in numeric format")
			sys.exit(1)
		row = int(sys.argv[3])

	return file, row, column

def read_xlsx(file, row, column):
	data = []
	workbook = openpyxl.load_workbook(file, data_only=False)
	sheet = workbook.active
	for row in sheet.iter_rows(min_row=row, max_row=sheet.max_row, min_col=column, max_col=column):
		data.append(row[0].value)
	workbook.close()
	return data

def is_invalid_data(data):
	return (data is None or data == "미제출")

def calculate_final_score(data):
	for i in range(len(data)):
		if is_invalid_data(data[i]):
			continue
		data[i] = str(data[i])
		if (data[i][0] == '='):
			data[i] = '(' + str(data[i][1:]) + ')'
		data[i] = "=30 + 0.7 * " + data[i]
	return data

def save_xlsx(file, data, row, column):
	try:
		workbook = openpyxl.load_workbook(file)
		sheet = workbook.active
		for i in range(len(data)):
			sheet.cell(row=row + i, column=column, value=data[i])
		workbook.save(file)
		workbook.close()
	except Exception as e:
		print("Error: " + str(e))
		sys.exit(1)

def process(file, row, column):
	data = read_xlsx(file, row, column)
	data = calculate_final_score(data)
	save_xlsx(file, data, row, column)

if __name__ == "__main__":
	process(*process_arguments())